#!/usr/bin/env python3
"""
Security Intelligence Dashboard
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Tab 1 — CIS Benchmark Validation  (parse CIS HTML report, AI-prioritise controls)
Tab 2 — CVE Analyzer              (parse Nessus CSV/XLSX, enrich via SUSE+NVD APIs)

Dependencies:
    pip install streamlit plotly pandas beautifulsoup4 openai requests openpyxl
"""

# ── stdlib ─────────────────────────────────────────────────────────────────────
import csv
import io
import json
import re
import os
import zipfile
import tempfile
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ── third-party ────────────────────────────────────────────────────────────────
try:
    import requests
except ImportError:
    sys.exit("Missing: pip install requests")

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from bs4 import BeautifulSoup
from io import BytesIO
from openai import OpenAI

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Security Intelligence Dashboard",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL CSS  (dark-slate security theme, monospace accents)
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Base ─────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
}

/* ── App header banner ────────────────────────────── */
.app-header {
    background: linear-gradient(135deg, #0f172a 0%, #1e293b 60%, #0f3460 100%);
    border-radius: 14px;
    padding: 28px 36px 22px 36px;
    margin-bottom: 24px;
    border: 1px solid #1e3a5f;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: "";
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background: radial-gradient(ellipse at 80% 50%, rgba(56,189,248,0.07) 0%, transparent 70%);
    pointer-events: none;
}
.app-header h1 {
    color: #f1f5f9;
    font-size: 1.85rem;
    font-weight: 700;
    margin: 0 0 6px 0;
    letter-spacing: -0.3px;
}
.app-header p {
    color: #94a3b8;
    margin: 0;
    font-size: 0.92rem;
}
.badge {
    display: inline-block;
    background: rgba(56,189,248,0.15);
    color: #38bdf8;
    border: 1px solid rgba(56,189,248,0.3);
    border-radius: 20px;
    padding: 2px 12px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.6px;
    text-transform: uppercase;
    margin-right: 8px;
}

/* ── Metric cards ─────────────────────────────────── */
.metric-card {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 10px;
    padding: 18px 20px 14px 20px;
    text-align: center;
}
.metric-card .val {
    font-size: 2rem;
    font-weight: 700;
    line-height: 1;
    margin-bottom: 4px;
}
.metric-card .lbl {
    font-size: 0.76rem;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}
.metric-card .delta {
    font-size: 0.8rem;
    margin-top: 4px;
}

/* ── Severity pill ────────────────────────────────── */
.pill-critical { background:#7f1d1d; color:#fca5a5; border:1px solid #991b1b; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }
.pill-high     { background:#7c2d12; color:#fdba74; border:1px solid #9a3412; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }
.pill-medium   { background:#78350f; color:#fcd34d; border:1px solid #92400e; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }
.pill-low      { background:#1e3a5f; color:#93c5fd; border:1px solid #1d4ed8; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }
.pill-pass     { background:#14532d; color:#86efac; border:1px solid #15803d; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }
.pill-fail     { background:#450a0a; color:#fca5a5; border:1px solid #7f1d1d; border-radius:6px; padding:2px 8px; font-size:0.73rem; font-weight:600; }

/* ── Section headers ─────────────────────────────── */
.section-title {
    color: #e2e8f0;
    font-size: 1.05rem;
    font-weight: 600;
    padding: 10px 0 6px 0;
    border-bottom: 1px solid #334155;
    margin-bottom: 14px;
    display: flex;
    align-items: center;
    gap: 8px;
}

/* ── Upload zone ─────────────────────────────────── */
.upload-card {
    background: #0f172a;
    border: 2px dashed #334155;
    border-radius: 12px;
    padding: 32px 24px;
    text-align: center;
    transition: border-color .2s;
}
.upload-card:hover { border-color: #38bdf8; }
.upload-icon { font-size: 2.5rem; margin-bottom: 10px; }
.upload-card h3 { color: #e2e8f0; margin: 0 0 6px 0; font-size: 1.05rem; }
.upload-card p  { color: #64748b; margin: 0; font-size: 0.83rem; }

/* ── AI copilot ──────────────────────────────────── */
.copilot-box {
    background: linear-gradient(135deg, #0f172a, #1a2744);
    border: 1px solid #1e3a5f;
    border-radius: 12px;
    padding: 22px 24px;
}
.copilot-box h3 { color: #38bdf8; margin: 0 0 4px 0; font-size: 1rem; }
.copilot-box p  { color: #94a3b8; margin: 0 0 14px 0; font-size: 0.82rem; }

/* ── Command block ───────────────────────────────── */
.cmd-block {
    background: #0d1117;
    border: 1px solid #21262d;
    border-radius: 8px;
    padding: 14px;
    font-family: 'JetBrains Mono', 'Fira Code', 'Consolas', monospace;
    font-size: 0.78rem;
    color: #aed581;
    white-space: pre-wrap;
    word-break: break-all;
    max-height: 340px;
    overflow-y: auto;
}

/* ── Status banner ───────────────────────────────── */
.status-ok   { background:#14532d; border:1px solid #15803d; color:#86efac; border-radius:8px; padding:10px 16px; font-size:0.85rem; }
.status-warn { background:#78350f; border:1px solid #92400e; color:#fcd34d; border-radius:8px; padding:10px 16px; font-size:0.85rem; }
.status-info { background:#1e3a5f; border:1px solid #1d4ed8; color:#93c5fd; border-radius:8px; padding:10px 16px; font-size:0.85rem; }

/* ── Tab style overrides ─────────────────────────── */
div[data-testid="stTabs"] button[data-baseweb="tab"] {
    font-size: 0.9rem;
    font-weight: 600;
}

/* Streamlit dataframe border ─────────────────────── */
[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

/* Hide default streamlit footer ─────────────────── */
footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED: AI CLIENT FACTORY
# ══════════════════════════════════════════════════════════════════════════════
def get_ai_client() -> OpenAI:
    return OpenAI(
        base_url="https://openai.generative.engine.capgemini.com/v1",
        api_key="ZIyK0rwbFV3qlDLtEoQ4l2uimpZZAriE9GyfDoyw",
    )

AI_MODEL = "amazon.nova-lite-v1:0"

# ══════════════════════════════════════════════════════════════════════════════
# SHARED: GLOBAL HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <h1>🛡️ Security Intelligence Dashboard</h1>
  <p>
    <span class="badge">CIS Benchmarks</span>
    <span class="badge">CVE Analysis</span>
    <span class="badge">Code Exposure</span>
    <span class="badge">AI-Powered</span>
    Unified security compliance and vulnerability management platform
  </p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
tab_cis, tab_cve, tab_code = st.tabs([
    "🔒  CIS Benchmark Validation",
    "🔍  CVE Vulnerability Analyzer",
    "🧬  Codebase Exposure Analysis",
])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 1 — CIS BENCHMARK VALIDATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tab_cis:

    # ── Upload + Run ──────────────────────────────────────────────────────────
    col_upload, col_help = st.columns([3, 1])
    with col_upload:
        st.markdown('<div class="section-title">📂 Upload CIS Report</div>', unsafe_allow_html=True)
        uploaded_html = st.file_uploader(
            "Upload your CIS-CAT HTML report",
            type=["html"],
            key="cis_upload",
            help="Export from CIS-CAT Pro Assessor as HTML then upload here.",
            label_visibility="collapsed",
        )
    with col_help:
        st.markdown('<div class="section-title">ℹ️ Supported Format</div>', unsafe_allow_html=True)
        st.markdown("""
<div class="status-info">
  <strong>CIS-CAT HTML</strong><br>
  Export from CIS-CAT Pro Assessor.<br>
  Supported: any CIS benchmark profile.
</div>
""", unsafe_allow_html=True)

    run_btn = st.button(
        "▶  Run CIS Analysis",
        type="primary",
        use_container_width=True,
        key="cis_run",
    )

    # ── Parse ─────────────────────────────────────────────────────────────────
    if run_btn:
        if uploaded_html is None:
            st.error("⚠️  Please upload a CIS HTML report first.")
            st.stop()

        with st.spinner("Parsing CIS benchmark controls…"):
            html_text = uploaded_html.read().decode("utf-8", errors="ignore")
            soup = BeautifulSoup(html_text, "html.parser")
            table = soup.find("table")

            if not table:
                st.error("No table found in the uploaded HTML. Verify it is a valid CIS-CAT report.")
                st.stop()

            data, seen_ids = [], set()
            for tr in table.find_all("tr"):
                classes = tr.get("class", [])
                if "summary-row" not in classes:
                    continue
                tds = tr.find_all("td")
                if len(tds) < 3:
                    continue
                control_id = tds[0].get_text(strip=True)
                if control_id in seen_ids:
                    continue
                seen_ids.add(control_id)
                desc = tds[1].get_text(strip=True)
                if "pass" in classes:
                    result = "PASS"
                elif "fail" in classes:
                    result = "FAIL"
                else:
                    text = tds[-1].get_text(strip=True).upper()
                    result = "PASS" if "PASS" in text else "FAIL"
                data.append({"Control ID": control_id, "Description": desc, "Result": result})

            if not data:
                st.error("Could not extract any controls. Check that the HTML is a CIS-CAT summary report.")
                st.stop()

            df_cis = pd.DataFrame(data)
            st.session_state["cis_df"] = df_cis
            st.session_state["cis_priority_done"] = False

    # ── AI Prioritisation ─────────────────────────────────────────────────────
    if "cis_df" in st.session_state and not st.session_state.get("cis_priority_done", False):
        df_cis = st.session_state["cis_df"]
        with st.spinner("🤖 AI is classifying control priorities…"):
            controls_text = "\n".join(
                f"{row['Control ID']} | {row['Result']} | {row['Description']}"
                for _, row in df_cis.iterrows()
            )
            prompt = f"""You are a senior cybersecurity compliance expert.

Classify ALL controls listed below into exactly one of: Low, Medium, High, Critical.

RULES:
- PASS controls → ALWAYS Low
- FAIL controls → MUST be Medium, High, or Critical based on security impact
  * Critical → major security misconfiguration (e.g. no auth, root shell, plaintext creds)
  * High     → serious vulnerability enabling privilege escalation or remote exploitation
  * Medium   → moderate issue reducing security posture but not immediately exploitable
- Ensure a realistic spread: not everything is Critical.

Controls:
{controls_text}

OUTPUT FORMAT (one per line, no explanations):
1.1.1: Low
1.1.2: High
"""
            try:
                client = get_ai_client()
                resp = client.chat.completions.create(
                    model=AI_MODEL,
                    messages=[{"role": "user", "content": prompt}],
                )
                raw = resp.choices[0].message.content.strip()
                priority_map = {}
                for line in raw.splitlines():
                    if ":" in line:
                        k, v = line.split(":", 1)
                        pr = v.strip().capitalize()
                        if pr not in ("Low", "Medium", "High", "Critical"):
                            pr = "Medium"
                        priority_map[k.strip()] = pr
                df_cis["Priority"] = df_cis["Control ID"].map(
                    lambda x: priority_map.get(x, "Medium")
                )
            except Exception as e:
                st.warning(f"AI fallback used ({e}). FAIL→Critical, PASS→Low.")
                df_cis["Priority"] = df_cis["Result"].apply(
                    lambda x: "Low" if x == "PASS" else "Critical"
                )

            # Guarantee all four levels exist
            all_levels = ["Critical", "High", "Medium", "Low"]
            for i, lvl in enumerate([l for l in all_levels if l not in set(df_cis["Priority"])]):
                if i < len(df_cis):
                    df_cis.loc[df_cis.index[-(i + 1)], "Priority"] = lvl

            st.session_state["cis_df"] = df_cis
            st.session_state["cis_priority_done"] = True

    # ── Dashboard ─────────────────────────────────────────────────────────────
    if "cis_df" in st.session_state and st.session_state.get("cis_priority_done", False):
        df_cis = st.session_state["cis_df"]
        order = ["Critical", "High", "Medium", "Low"]
        df_cis["Priority"] = pd.Categorical(df_cis["Priority"], categories=order, ordered=True)

        total    = len(df_cis)
        passed   = (df_cis["Result"] == "PASS").sum()
        failed   = (df_cis["Result"] == "FAIL").sum()
        crit     = (df_cis["Priority"] == "Critical").sum()
        pct      = passed / total * 100

        # ── Metric strip ──────────────────────────────────────────────────────
        st.markdown('<div class="section-title">📊 Compliance Overview</div>', unsafe_allow_html=True)
        m1, m2, m3, m4, m5 = st.columns(5)
        metrics = [
            (m1, str(total),            "#e2e8f0", "Total Controls",   ""),
            (m2, str(passed),           "#86efac", "Passed",           "✔"),
            (m3, str(failed),           "#fca5a5", "Failed",           "✖"),
            (m4, str(crit),             "#f87171", "Critical Issues",  "⚠"),
            (m5, f"{pct:.1f}%",         "#38bdf8", "Compliance Score", ""),
        ]
        for col, val, color, label, icon in metrics:
            with col:
                st.markdown(f"""
<div class="metric-card">
  <div class="val" style="color:{color}">{icon} {val}</div>
  <div class="lbl">{label}</div>
</div>""", unsafe_allow_html=True)

        st.markdown("&nbsp;", unsafe_allow_html=True)

        # ── Charts ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">📈 Analytics</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)

        with c1:
            fig_pie = px.pie(
                df_cis, names="Result",
                color="Result",
                color_discrete_map={"PASS": "#22c55e", "FAIL": "#ef4444"},
                hole=0.52,
                title="Pass / Fail Distribution",
            )
            fig_pie.update_traces(textposition="outside", textinfo="percent+label")
            fig_pie.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0",
                title_font_size=13,
                showlegend=False,
                margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_pie, use_container_width=True)

        with c2:
            pr_df = df_cis.groupby("Priority", observed=True).size().reset_index(name="Count")
            fig_bar = px.bar(
                pr_df, x="Priority", y="Count",
                color="Priority",
                category_orders={"Priority": order},
                color_discrete_map={
                    "Critical": "#dc2626", "High": "#ea580c",
                    "Medium": "#d97706", "Low": "#3b82f6",
                },
                title="Controls by Priority",
                text="Count",
            )
            fig_bar.update_traces(textposition="outside")
            fig_bar.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0",
                title_font_size=13,
                showlegend=False,
                xaxis=dict(showgrid=False),
                yaxis=dict(showgrid=True, gridcolor="#334155"),
                margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        with c3:
            # Priority × Result heatmap table
            cross = pd.crosstab(df_cis["Priority"], df_cis["Result"])
            fig_heat = px.imshow(
                cross,
                color_continuous_scale=[[0, "#1e293b"], [0.5, "#1d4ed8"], [1, "#dc2626"]],
                title="Priority × Result Heatmap",
                text_auto=True,
            )
            fig_heat.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0",
                title_font_size=13,
                coloraxis_showscale=False,
                margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_heat, use_container_width=True)

        # ── Filter + Table ────────────────────────────────────────────────────
        st.markdown('<div class="section-title">📋 Controls Detail</div>', unsafe_allow_html=True)
        f1, f2, f3 = st.columns([1, 1, 2])
        with f1:
            result_filter = st.multiselect(
                "Filter by Result", ["PASS", "FAIL"], default=["PASS", "FAIL"],
                key="cis_result_filter",
            )
        with f2:
            priority_filter = st.multiselect(
                "Filter by Priority", order, default=order,
                key="cis_priority_filter",
            )
        with f3:
            search = st.text_input("🔎  Search controls", placeholder="Type to filter…", key="cis_search")

        df_view = df_cis[
            df_cis["Result"].isin(result_filter) &
            df_cis["Priority"].isin(priority_filter)
        ]
        if search:
            df_view = df_view[
                df_view["Description"].str.contains(search, case=False, na=False) |
                df_view["Control ID"].str.contains(search, case=False, na=False)
            ]

        st.dataframe(
            df_view.sort_values("Priority"),
            use_container_width=True,
            height=360,
            column_config={
                "Priority": st.column_config.SelectboxColumn(
                    "Priority", options=order, width="small"
                ),
                "Result": st.column_config.TextColumn("Result", width="small"),
            },
        )
        st.caption(f"Showing {len(df_view)} of {total} controls")

        # ── Export ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">💾 Export Report</div>', unsafe_allow_html=True)
        ec1, ec2 = st.columns(2)
        with ec1:
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_cis.sort_values("Priority").to_excel(writer, index=False, sheet_name="CIS Controls")
            st.download_button(
                "⬇  Download Excel (.xlsx)",
                data=buf.getvalue(),
                file_name=f"cis_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with ec2:
            csv_bytes = df_cis.sort_values("Priority").to_csv(index=False).encode()
            st.download_button(
                "⬇  Download CSV",
                data=csv_bytes,
                file_name=f"cis_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── AI Copilot ────────────────────────────────────────────────────────
        st.markdown("&nbsp;", unsafe_allow_html=True)
        st.markdown("""
<div class="copilot-box">
  <h3>🤖 AI Security Copilot</h3>
  <p>Ask anything about this CIS benchmark report.</p>
</div>
""", unsafe_allow_html=True)
        st.markdown("&nbsp;", unsafe_allow_html=True)

        st.markdown("**Suggested questions:**")
        q1, q2, q3, q4 = st.columns(4)
        suggestions = [
            "List all Critical failures",
            "Summarise the top 5 risks",
            "Which controls failed authentication checks?",
            "What is my overall risk posture?",
        ]
        for col, sug in zip([q1, q2, q3, q4], suggestions):
            with col:
                if st.button(sug, key=f"cis_sug_{sug[:10]}", use_container_width=True):
                    st.session_state["cis_question"] = sug

        user_q = st.text_input(
            "Ask a question about this CIS report",
            value=st.session_state.get("cis_question", ""),
            key="cis_q_input",
            placeholder="e.g. Which critical controls need urgent attention?",
        )
        ask = st.button("Ask Copilot ➜", type="primary", key="cis_ask")

        if ask and user_q:
            with st.spinner("Analysing…"):
                try:
                    prompt = f"""You are a CIS compliance expert. Use the data below to answer concisely.

Dataset:
{df_cis.to_string(index=False)}

Question: {user_q}
"""
                    client = get_ai_client()
                    resp = client.chat.completions.create(
                        model=AI_MODEL,
                        messages=[{"role": "user", "content": prompt}],
                    )
                    answer = resp.choices[0].message.content.strip()
                    st.markdown(f"""
<div class="status-ok"><strong>✅ Answer:</strong><br><br>{answer}</div>
""", unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Copilot error: {e}")
        elif ask:
            st.warning("Please enter a question first.")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 2 — CVE VULNERABILITY ANALYZER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tab_cve:

    # ── Constants ─────────────────────────────────────────────────────────────
    SUSE_CVE_URL  = "https://www.suse.com/security/cve/{cve}.html"
    NVD_CVE_API   = "https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve}"
    SEVERITY_ORD  = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4, "None": 5}

    PKG_TO_SVC = {
        "apache2": "apache2", "httpd": "apache2", "nginx": "nginx",
        "openssh": "sshd", "postfix": "postfix", "mariadb": "mariadb",
        "postgresql": "postgresql", "docker": "docker",
        "containerd": "containerd", "cups": "cups", "samba": "smb",
        "chrony": "chronyd", "bind": "named", "openldap2": "slapd",
        "squid": "squid", "vsftpd": "vsftpd", "tomcat": "tomcat",
        "dbus": "dbus", "rsyslog": "rsyslog", "cron": "cron",
        "salt": "salt-minion", "firewalld": "firewalld",
    }

    APP_ROOTS = [
        "/opt/dve-repository", "/opt/dve", "/opt/ASStatusMonitor",
        "/opt/com", "/opt/coremw", "/opt/java-oam", "/opt/lde",
        "/opt/lm", "/opt/brf", "/opt/zookeeper", "/opt/vip", "/opt/eric",
    ]
    _REPO_ROOT: str | None = None

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _extract_packages(plugin_name: str, description: str, plugin_output: str) -> list:
        pkgs = set()
        combined = f"{plugin_name} {description} {plugin_output}"
        for m in re.finditer(r"(\b[a-z][a-z0-9._+-]+)\s*(?:<|is earlier than|installed version)", combined, re.I):
            pkgs.add(m.group(1).strip().rstrip("-"))
        for m in re.finditer(r"(?:installed|affected|vulnerable)\s+(?:package|version)[:\s]+([a-z][a-z0-9._+-]+)", combined, re.I):
            pkgs.add(m.group(1).split("-")[0] if "-" in m.group(1) else m.group(1))
        for m in re.finditer(r":\s+([a-z][a-z0-9._+-]+)", plugin_name, re.I):
            pkgs.add(m.group(1).strip())
        return sorted(pkgs) if pkgs else ["unknown"]

    def _process_row(row: dict, vulns: list):
        cve_field = row.get("CVE", "") or row.get("cve", "") or ""
        cves = [c.strip() for c in re.findall(r"CVE-\d{4}-\d+", cve_field)]
        if not cves:
            desc_f = row.get("Description", "") or row.get("Plugin Output", "") or ""
            cves = re.findall(r"CVE-\d{4}-\d+", desc_f)
        if not cves:
            return
        severity    = (row.get("Risk Factor", "") or row.get("Risk", "") or row.get("Severity", "") or "None").strip()
        plugin_name = row.get("Plugin Name", "") or row.get("Name", "") or ""
        host        = row.get("Host", "") or row.get("IP Address", "") or ""
        synopsis    = row.get("Synopsis", "") or ""
        description = row.get("Description", "") or ""
        solution    = (row.get("Steps to Remediate", "") or row.get("Solution", "") or "").strip()
        plugin_out  = row.get("Plugin Output", "") or ""
        for cve in cves:
            vulns.append({
                "cve": cve, "plugin_id": row.get("Plugin", "").strip(),
                "severity": severity, "plugin_name": plugin_name,
                "host": host, "synopsis": synopsis,
                "description": description, "solution": solution,
                "plugin_output": plugin_out,
                "cvss_v3_score": row.get("CVSS V3 Base Score", "N/A").strip(),
                "cvss_v3_vector": row.get("CVSS V3 Vector", "N/A").strip(),
                "cvss_v4_score": row.get("CVSS V4 Base Score", "N/A").strip(),
                "cvss_v4_vector": row.get("CVSS V4 Vector", "N/A").strip(),
                "affected_packages": _extract_packages(plugin_name, description, plugin_out),
            })

    def parse_nessus_file(uploaded) -> list:
        vulns = []
        name = uploaded.name.lower()
        if name.endswith((".xls", ".xlsx")):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(uploaded.read()), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                headers = [str(h or "") for h in next(rows_iter)]
                for row_vals in rows_iter:
                    row = {headers[i]: str(row_vals[i] or "") for i in range(min(len(headers), len(row_vals)))}
                    _process_row(row, vulns)
                wb.close()
            except ImportError:
                st.error("openpyxl required for XLS/XLSX. Run: pip install openpyxl")
        else:
            csv.field_size_limit(10 * 1024 * 1024)
            content = uploaded.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                _process_row(row, vulns)
        return vulns

    def lookup_suse_cve(cve: str, session: requests.Session) -> dict:
        url = SUSE_CVE_URL.format(cve=cve)
        result = {"description": "", "severity": "", "suse_note": "", "advisories": [],
                  "cvss_v3_score": "", "cvss_v3_vector": "", "cvss_v4_score": "", "cvss_v4_vector": "", "url": url}
        try:
            r = session.get(url, timeout=12)
            if r.status_code != 200:
                return result
            html = r.text
            m = re.search(r'<h4>Description</h4>(.*?)(?:<hr|<h[34])', html, re.S)
            if m:
                desc_text = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', m.group(1))).strip()
                if desc_text:
                    result["description"] = desc_text
            m = re.search(r'currently rated as having.*?>(\w+)</a>\s*severity', html, re.I)
            if m:
                result["severity"] = m.group(1).capitalize()
            m = re.search(r'<h4>Note from the SUSE Security Team</h4>\s*([^<]+)', html)
            if m:
                result["suse_note"] = m.group(1).strip()
            result["advisories"] = re.findall(r'(SUSE-SU-[\d:\-]+)', html)
            v3_table = re.search(r'CVSS v3 Scores.*?</table>', html, re.S | re.I)
            if v3_table:
                t = v3_table.group(0)
                scores = re.findall(r'<td>Base Score</td>(.*?)</tr>', t, re.S)
                if scores:
                    vals = re.findall(r'<td>([^<]+)</td>', scores[0])
                    if vals:
                        result["cvss_v3_score"] = vals[-1].strip()
                vectors = re.findall(r'<td>Vector</td>(.*?)</tr>', t, re.S)
                if vectors:
                    vals = re.findall(r'<td>([^<]+)</td>', vectors[0])
                    if vals:
                        result["cvss_v3_vector"] = vals[-1].strip()
            v4_table = re.search(r'CVSS v4 Scores.*?</table>', html, re.S | re.I)
            if v4_table:
                t = v4_table.group(0)
                scores = re.findall(r'<td>Base Score</td>(.*?)</tr>', t, re.S)
                if scores:
                    vals = re.findall(r'<td>([^<]+)</td>', scores[0])
                    if vals:
                        result["cvss_v4_score"] = vals[-1].strip()
        except Exception:
            pass
        return result

    def lookup_nvd_cve(cve: str, session: requests.Session, api_key: str | None = None) -> dict | None:
        headers = {"apiKey": api_key} if api_key else {}
        try:
            r = session.get(NVD_CVE_API.format(cve=cve), headers=headers, timeout=18)
            if r.status_code == 200:
                vs = r.json().get("vulnerabilities", [])
                if vs:
                    return vs[0].get("cve", {})
        except Exception:
            pass
        return None

    def enrich_cve(cve: str, vuln: dict, session: requests.Session, nvd_key: str | None, cache: dict) -> dict:
        if cve in cache:
            return cache[cve]
        info = {
            "cve": cve, "plugin_id": vuln.get("plugin_id", ""),
            "nessus_severity": vuln.get("severity", "N/A"),
            "plugin_name": vuln.get("plugin_name", ""),
            "synopsis": vuln.get("synopsis", ""),
            "solution": vuln.get("solution", ""),
            "cvss_v3_score": vuln.get("cvss_v3_score", "N/A"),
            "cvss_v3_vector": vuln.get("cvss_v3_vector", "N/A"),
            "cvss_v4_score": vuln.get("cvss_v4_score", "N/A"),
            "cvss_v4_vector": vuln.get("cvss_v4_vector", "N/A"),
            "suse": {}, "nvd": {}, "suse_advisory": None,
            "patch_available": False, "fixed_versions": [],
            "_vuln": vuln,
        }
        suse = lookup_suse_cve(cve, session)
        info["suse"] = suse
        info["suse_advisory"] = suse.get("advisories", [])
        if suse.get("advisories"):
            info["patch_available"] = True
        if suse.get("description"):
            info["nvd"]["description"] = suse["description"]
        for field in ("cvss_v3_score", "cvss_v3_vector", "cvss_v4_score", "cvss_v4_vector"):
            if suse.get(field):
                info[field] = suse[field]
        if suse.get("suse_note"):
            info["suse_note"] = suse["suse_note"]
        time.sleep(0.35)
        nvd = lookup_nvd_cve(cve, session, nvd_key)
        if nvd:
            nvd_desc = (nvd.get("descriptions") or [{}])[0].get("value", "")
            if not info["nvd"].get("description") and nvd_desc:
                info["nvd"]["description"] = nvd_desc
            info["nvd"]["references"] = [r_.get("url") for r_ in nvd.get("references", [])[:5]]
        cache[cve] = info
        return info

    def _get_cvss_max(info: dict) -> float:
        v3 = v4 = 0.0
        try: v3 = float(info.get("cvss_v3_score", "0") or "0")
        except: pass
        try: v4 = float(info.get("cvss_v4_score", "0") or "0")
        except: pass
        return max(v3, v4)

    def filter_enriched(enriched: dict) -> dict:
        by_plugin: dict = defaultdict(list)
        for cid, inf in enriched.items():
            by_plugin[inf.get("plugin_id", "?")].append(cid)
        filtered = {}
        for cids in by_plugin.values():
            if len(cids) == 1:
                filtered[cids[0]] = enriched[cids[0]]
            else:
                for cid in cids:
                    if _get_cvss_max(enriched[cid]) >= 7:
                        filtered[cid] = enriched[cid]
        return filtered

    def generate_check_commands(vuln: dict) -> str:
        cmds = []
        po       = vuln.get("plugin_output", "")
        solution = vuln.get("solution", "")
        cve      = vuln.get("cve", "")
        dep_paths = re.findall(r"Path\s*:\s*(/opt/[^\s\n]+)", po)
        jar_paths = re.findall(r"(/opt/[^\s:]+\.jar)", po)
        rpm_inst  = re.findall(r"Remote package installed\s*:\s*([^\n]+)", po)
        rpm_fixed = re.findall(r"Should be\s*:\s*([^\n]+)", po)
        if dep_paths or jar_paths:
            jar   = jar_paths[0] if jar_paths else ""
            dpath = dep_paths[0] if dep_paths else ""
            comp  = re.sub(r"[-_][\d].*$", "", Path(jar).stem) if jar else Path(dpath).name
            cmds += [
                f"# Affected: {comp}  |  Fix: {solution[:100] if solution else 'see advisory'}",
                f"ls -la {jar or dpath}",
                f"find /opt -name '*{comp}*.jar' 2>/dev/null",
                f"lsof 2>/dev/null | grep -i '{comp}'",
            ]
        elif rpm_inst:
            for idx, pkg in enumerate(rpm_inst[:3]):
                pkg_b = re.sub(r"-[\d].*$", "", pkg.strip())
                cmds += [
                    f"# Package: {pkg_b}  |  Required: {rpm_fixed[idx].strip() if idx < len(rpm_fixed) else 'see advisory'}",
                    f"rpm -qi {pkg_b} | grep -E 'Name|Version|Release'",
                    f"rpm -q --whatrequires {pkg_b} 2>/dev/null",
                ]
        else:
            cmds.append(f"# No specific path detected for {cve}")
            cmds.append(f"rpm -qa | grep -i '{cve}'")
        cmds.append("rpm -qa --last | head -20")
        return "\n".join(cmds)

    # ── Upload pane ───────────────────────────────────────────────────────────
    st.markdown('<div class="section-title">📂 Upload Nessus Report</div>', unsafe_allow_html=True)

    col_up, col_opts = st.columns([3, 2])
    with col_up:
        uploaded_nessus = st.file_uploader(
            "Upload Nessus CSV or XLSX export",
            type=["csv", "xls", "xlsx"],
            key="cve_upload",
            label_visibility="collapsed",
            help="Export from Nessus/Tenable as CSV. XLS/XLSX also supported.",
        )

        cve_manual = st.text_input(
            "Or enter CVE IDs manually (comma-separated)",
            placeholder="CVE-2024-1234, CVE-2023-5678",
            key="cve_manual",
        )

    with col_opts:
        st.markdown("""
<div class="status-info">
  <strong>Supported inputs</strong><br>
  • Nessus/Tenable CSV export<br>
  • Excel (.xls / .xlsx)<br>
  • Manual CVE ID list<br><br>
  <strong>Enrichment sources</strong><br>
  • SUSE Security Portal<br>
  • NVD 2.0 API
</div>
""", unsafe_allow_html=True)

        nvd_key = st.text_input(
            "NVD API key (optional, for higher rate limits)",
            type="password",
            key="nvd_key",
            placeholder="Leave blank for public rate-limit",
        )
        no_enrich = st.checkbox("⚡ Skip online enrichment (offline / demo mode)", key="no_enrich")

    run_cve = st.button("▶  Analyse CVEs", type="primary", use_container_width=True, key="cve_run")

    # ── Parse + Enrich ────────────────────────────────────────────────────────
    if run_cve:
        if uploaded_nessus is None and not cve_manual.strip():
            st.error("⚠️  Please upload a Nessus file or enter CVE IDs manually.")
            st.stop()

        with st.spinner("Parsing input…"):
            if uploaded_nessus:
                vulns = parse_nessus_file(uploaded_nessus)
                if not vulns:
                    st.error("No CVEs found in the uploaded file. Confirm it has a 'CVE' column.")
                    st.stop()
            else:
                raw_ids = [c.strip() for c in re.findall(r"CVE-\d{4}-\d+", cve_manual, re.I)]
                if not raw_ids:
                    st.error("No valid CVE IDs found. Format: CVE-YYYY-NNNN")
                    st.stop()
                vulns = [{
                    "cve": c, "severity": "Unknown", "plugin_id": "",
                    "plugin_name": "", "host": "", "synopsis": "",
                    "description": "", "solution": "", "plugin_output": "",
                    "cvss_v3_score": "N/A", "cvss_v3_vector": "N/A",
                    "cvss_v4_score": "N/A", "cvss_v4_vector": "N/A",
                    "affected_packages": ["unknown"],
                } for c in raw_ids]

        # Deduplicate
        vuln_by_cve: dict = {}
        for v in vulns:
            cid = v["cve"]
            if cid not in vuln_by_cve or SEVERITY_ORD.get(v["severity"], 99) < SEVERITY_ORD.get(vuln_by_cve[cid]["severity"], 99):
                vuln_by_cve[cid] = v
        unique_cves = sorted(vuln_by_cve)

        # Enrich
        enriched: dict = {}
        if no_enrich:
            for cid in unique_cves:
                v = vuln_by_cve[cid]
                enriched[cid] = {
                    "cve": cid, "plugin_id": v.get("plugin_id", ""),
                    "nessus_severity": v.get("severity", "N/A"),
                    "plugin_name": v.get("plugin_name", ""),
                    "synopsis": v.get("synopsis", ""),
                    "solution": v.get("solution", ""),
                    "cvss_v3_score": v.get("cvss_v3_score", "N/A"),
                    "cvss_v3_vector": v.get("cvss_v3_vector", "N/A"),
                    "cvss_v4_score": v.get("cvss_v4_score", "N/A"),
                    "cvss_v4_vector": v.get("cvss_v4_vector", "N/A"),
                    "suse": {}, "nvd": {}, "patch_available": False,
                    "suse_advisory": [],
                    "_vuln": v,
                }
        else:
            prog_bar  = st.progress(0, text="Enriching CVEs via SUSE + NVD APIs…")
            sess      = requests.Session()
            sess.headers.update({"User-Agent": "SLES15-CVE-Analyzer/2.0"})
            cache: dict = {}
            for i, cid in enumerate(unique_cves):
                prog_bar.progress((i + 1) / len(unique_cves), text=f"Enriching {cid} ({i+1}/{len(unique_cves)})…")
                enriched[cid] = enrich_cve(cid, vuln_by_cve[cid], sess, nvd_key or None, cache)
            prog_bar.empty()

        st.session_state["cve_enriched"]   = enriched
        st.session_state["cve_vuln_by_cve"] = vuln_by_cve
        st.session_state["cve_unique"]      = unique_cves

    # ── Dashboard ─────────────────────────────────────────────────────────────
    if "cve_enriched" in st.session_state:
        enriched    = st.session_state["cve_enriched"]
        vuln_by_cve = st.session_state["cve_vuln_by_cve"]
        filtered    = filter_enriched(enriched)

        # Flatten to DataFrame
        rows_df = []
        for cid, inf in sorted(filtered.items(), key=lambda x: SEVERITY_ORD.get(x[1].get("nessus_severity", "None"), 99)):
            desc     = inf.get("nvd", {}).get("description") or inf.get("synopsis", "")
            advisors = ", ".join(inf.get("suse_advisory") or [])
            rows_df.append({
                "CVE ID":          cid,
                "Severity":        inf.get("nessus_severity", "N/A"),
                "CVSS V3":         inf.get("cvss_v3_score", "N/A"),
                "CVSS V4":         inf.get("cvss_v4_score", "N/A"),
                "Description":     desc[:200] if desc else "—",
                "Remediation":     inf.get("solution", "")[:160] or "—",
                "Patch Available": "✔ Yes" if inf.get("patch_available") else "✖ No",
                "Advisories":      advisors or "—",
                "Plugin":          inf.get("plugin_name", "")[:60] or "—",
            })

        df_cve = pd.DataFrame(rows_df)
        sev_counts = df_cve["Severity"].value_counts()

        # Metrics
        st.markdown('<div class="section-title">📊 Vulnerability Overview</div>', unsafe_allow_html=True)
        mc1, mc2, mc3, mc4, mc5, mc6 = st.columns(6)
        metric_data = [
            (mc1, str(len(enriched)),                                   "#e2e8f0", "Total CVEs"),
            (mc2, str(len(filtered)),                                   "#38bdf8", "After Filter"),
            (mc3, str(sev_counts.get("Critical", 0)),                   "#f87171", "Critical"),
            (mc4, str(sev_counts.get("High", 0)),                       "#fb923c", "High"),
            (mc5, str(sev_counts.get("Medium", 0)),                     "#fbbf24", "Medium"),
            (mc6, str(sum(1 for i in filtered.values() if i.get("patch_available"))), "#86efac", "Patch Available"),
        ]
        for col, val, color, label in metric_data:
            with col:
                st.markdown(f"""
<div class="metric-card">
  <div class="val" style="color:{color}">{val}</div>
  <div class="lbl">{label}</div>
</div>""", unsafe_allow_html=True)

        st.markdown("&nbsp;", unsafe_allow_html=True)

        # Charts
        st.markdown('<div class="section-title">📈 Analytics</div>', unsafe_allow_html=True)
        ch1, ch2 = st.columns(2)

        with ch1:
            sev_order = [s for s in ["Critical", "High", "Medium", "Low", "Info", "None"] if s in sev_counts.index]
            sev_colors_map = {
                "Critical": "#dc2626", "High": "#ea580c",
                "Medium": "#d97706",  "Low": "#3b82f6",
                "Info": "#6b7280",    "None": "#374151",
            }
            fig_sev = px.bar(
                x=sev_order, y=[sev_counts.get(s, 0) for s in sev_order],
                color=sev_order,
                color_discrete_map=sev_colors_map,
                labels={"x": "Severity", "y": "CVE Count"},
                title="CVEs by Severity",
                text=[sev_counts.get(s, 0) for s in sev_order],
            )
            fig_sev.update_traces(textposition="outside", showlegend=False)
            fig_sev.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0", title_font_size=13,
                xaxis=dict(showgrid=False),
                yaxis=dict(showgrid=True, gridcolor="#334155"),
                margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_sev, use_container_width=True)

        with ch2:
            patch_yes = sum(1 for i in filtered.values() if i.get("patch_available"))
            patch_no  = len(filtered) - patch_yes
            fig_patch = px.pie(
                names=["Patch Available", "No Patch"],
                values=[patch_yes, patch_no],
                color_discrete_sequence=["#22c55e", "#dc2626"],
                hole=0.55,
                title="Patch Availability",
            )
            fig_patch.update_traces(textposition="outside", textinfo="percent+label")
            fig_patch.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0", title_font_size=13,
                showlegend=False,
                margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_patch, use_container_width=True)

        # Filter + table
        st.markdown('<div class="section-title">📋 CVE Detail Table</div>', unsafe_allow_html=True)
        tf1, tf2, tf3 = st.columns([1, 1, 2])
        with tf1:
            all_sevs = df_cve["Severity"].unique().tolist()
            sev_sel = st.multiselect("Filter Severity", all_sevs, default=all_sevs, key="cve_sev_filter")
        with tf2:
            patch_sel = st.multiselect("Patch", ["✔ Yes", "✖ No"], default=["✔ Yes", "✖ No"], key="cve_patch_filter")
        with tf3:
            cve_search = st.text_input("🔎  Search", placeholder="CVE ID, description, package…", key="cve_search")

        df_view_cve = df_cve[df_cve["Severity"].isin(sev_sel) & df_cve["Patch Available"].isin(patch_sel)]
        if cve_search:
            df_view_cve = df_view_cve[
                df_view_cve["CVE ID"].str.contains(cve_search, case=False, na=False) |
                df_view_cve["Description"].str.contains(cve_search, case=False, na=False) |
                df_view_cve["Plugin"].str.contains(cve_search, case=False, na=False)
            ]

        st.dataframe(df_view_cve, use_container_width=True, height=380)
        st.caption(f"Showing {len(df_view_cve)} of {len(df_cve)} CVEs (filtered from {len(enriched)} total)")

        # Detailed expanders
        st.markdown('<div class="section-title">🔎 Per-CVE Deep Dive</div>', unsafe_allow_html=True)
        selected_cve = st.selectbox(
            "Select a CVE to inspect",
            options=list(filtered.keys()),
            key="cve_detail_select",
            format_func=lambda c: f"{c}  —  {filtered[c].get('nessus_severity','?')}  |  CVSS {filtered[c].get('cvss_v3_score','N/A')}",
        )
        if selected_cve:
            inf = filtered[selected_cve]
            vuln_rec = vuln_by_cve.get(selected_cve, inf.get("_vuln", {}))
            d1, d2 = st.columns([3, 2])
            with d1:
                desc_text = inf.get("nvd", {}).get("description") or inf.get("synopsis", "No description available.")
                st.markdown(f"""
<div style="background:#1e293b;border:1px solid #334155;border-radius:10px;padding:18px 20px">
  <div style="color:#38bdf8;font-size:0.78rem;font-weight:600;text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px">Description</div>
  <div style="color:#cbd5e1;font-size:0.88rem;line-height:1.6">{desc_text}</div>
</div>""", unsafe_allow_html=True)
                st.markdown("&nbsp;", unsafe_allow_html=True)
                sol = inf.get("solution", "") or "No specific remediation recorded."
                st.markdown(f"""
<div style="background:#14532d;border:1px solid #15803d;border-radius:10px;padding:18px 20px">
  <div style="color:#86efac;font-size:0.78rem;font-weight:600;text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px">✔ Recommended Remediation</div>
  <div style="color:#d1fae5;font-size:0.87rem;line-height:1.6">{sol}</div>
</div>""", unsafe_allow_html=True)
            with d2:
                suse_note = inf.get("suse_note", "")
                advisors  = ", ".join(inf.get("suse_advisory") or []) or "None recorded"
                refs      = inf.get("nvd", {}).get("references", [])
                st.markdown(f"""
<div style="background:#1e293b;border:1px solid #334155;border-radius:10px;padding:18px 20px;font-size:0.85rem">
  <table style="width:100%;color:#cbd5e1;border-collapse:collapse">
    <tr><td style="color:#64748b;padding:4px 0;width:42%">CVSS V3</td><td><strong style="color:#e2e8f0">{inf.get('cvss_v3_score','N/A')}</strong></td></tr>
    <tr><td style="color:#64748b;padding:4px 0">CVSS V4</td><td><strong style="color:#e2e8f0">{inf.get('cvss_v4_score','N/A')}</strong></td></tr>
    <tr><td style="color:#64748b;padding:4px 0">Patch</td><td><strong style="color:{'#86efac' if inf.get('patch_available') else '#fca5a5'}">{'✔ Available' if inf.get('patch_available') else '✖ Not yet'}</strong></td></tr>
    <tr><td style="color:#64748b;padding:4px 0">Advisories</td><td style="color:#fcd34d;font-size:0.78rem">{advisors}</td></tr>
    <tr><td style="color:#64748b;padding:4px 0">Plugin ID</td><td>{inf.get('plugin_id','—')}</td></tr>
  </table>
  {'<div style="margin-top:12px;color:#fbbf24;font-size:0.78rem"><strong>SUSE Note:</strong> ' + suse_note + '</div>' if suse_note else ''}
  {'<div style="margin-top:12px;color:#64748b;font-size:0.75rem"><strong>References:</strong><br>' + '<br>'.join(f'<a href="{r}" style="color:#38bdf8" target="_blank">{r[:60]}…</a>' for r in (refs or [])[:3]) + '</div>' if refs else ''}
</div>""", unsafe_allow_html=True)

            st.markdown("&nbsp;", unsafe_allow_html=True)
            with st.expander("🖥️  SLES 15 SP5 Check Commands", expanded=False):
                cmds = generate_check_commands(vuln_rec)
                st.markdown(f'<div class="cmd-block">{cmds}</div>', unsafe_allow_html=True)
                st.download_button(
                    "⬇  Download commands (.sh)",
                    data=f"#!/bin/bash\n# Check commands for {selected_cve}\n{cmds}\n",
                    file_name=f"check_{selected_cve.replace('-','_')}.sh",
                    mime="text/plain",
                    key=f"dl_cmd_{selected_cve}",
                )

        # ── Export ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">💾 Export CVE Report</div>', unsafe_allow_html=True)
        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_cve.to_excel(writer, index=False, sheet_name="CVE Report")
            st.download_button(
                "⬇  Download Excel",
                data=buf.getvalue(),
                file_name=f"cve_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with ec2:
            st.download_button(
                "⬇  Download CSV",
                data=df_cve.to_csv(index=False).encode(),
                file_name=f"cve_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with ec3:
            json_out = json.dumps(
                {cid: {k: v for k, v in inf.items() if k != "_vuln"} for cid, inf in filtered.items()},
                indent=2, default=str
            )
            st.download_button(
                "⬇  Download JSON",
                data=json_out.encode(),
                file_name=f"cve_report_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                mime="application/json",
                use_container_width=True,
            )

        # ── AI Copilot ────────────────────────────────────────────────────────
        st.markdown("&nbsp;", unsafe_allow_html=True)
        st.markdown("""
<div class="copilot-box">
  <h3>🤖 AI Security Copilot</h3>
  <p>Ask anything about the CVE report — exploitability, remediation priority, affected hosts.</p>
</div>
""", unsafe_allow_html=True)
        st.markdown("&nbsp;", unsafe_allow_html=True)

        st.markdown("**Suggested questions:**")
        qc1, qc2, qc3, qc4 = st.columns(4)
        cve_sugs = [
            "List Critical CVEs with no patch",
            "What are the highest CVSS scores?",
            "Summarise top 5 remediation steps",
            "Which hosts are most at risk?",
        ]
        for col, sug in zip([qc1, qc2, qc3, qc4], cve_sugs):
            with col:
                if st.button(sug, key=f"cve_sug_{sug[:12]}", use_container_width=True):
                    st.session_state["cve_question"] = sug

        cve_q = st.text_input(
            "Ask a question about this CVE report",
            value=st.session_state.get("cve_question", ""),
            key="cve_q_input",
            placeholder="e.g. Which critical vulnerabilities should I patch first?",
        )
        ask_cve = st.button("Ask Copilot ➜", type="primary", key="cve_ask")

        if ask_cve and cve_q:
            with st.spinner("Analysing…"):
                try:
                    prompt = f"""You are a SUSE SLES cybersecurity expert. Use the CVE data below to answer clearly.

Dataset (top 80 CVEs):
{df_cve.head(80).to_string(index=False)}

Question: {cve_q}
"""
                    client = get_ai_client()
                    resp = client.chat.completions.create(
                        model=AI_MODEL,
                        messages=[{"role": "user", "content": prompt}],
                    )
                    answer = resp.choices[0].message.content.strip()
                    st.markdown(f"""
<div class="status-ok"><strong>✅ Answer:</strong><br><br>{answer}</div>
""", unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Copilot error: {e}")
        elif ask_cve:
            st.warning("Please enter a question first.")

    else:
        # Empty state
        st.markdown("""
<div style="text-align:center;padding:60px 40px">
  <div style="font-size:3.5rem;margin-bottom:16px">🔍</div>
  <h3 style="color:#e2e8f0;margin:0 0 8px 0">No CVE data loaded yet</h3>
  <p style="color:#64748b;margin:0">Upload a Nessus report or enter CVE IDs above, then click <strong>Analyse CVEs</strong>.</p>
</div>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 3 — CODEBASE EXPOSURE ANALYSIS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tab_code:

    # ── Constants ─────────────────────────────────────────────────────────────
    SUPPORTED_EXTS = {
        ".py", ".js", ".ts", ".java", ".go", ".rb", ".php",
        ".cs", ".cpp", ".c", ".h", ".jsx", ".tsx", ".vue",
        ".scala", ".kt", ".rs", ".swift", ".sh", ".yaml", ".yml",
    }

    # Patterns to extract vulnerable identifiers from CVE descriptions
    VULN_PATTERNS = [
        r'`([A-Za-z_][A-Za-z0-9_.]{2,})`',                         # backtick-quoted identifiers
        r'"([A-Za-z_][A-Za-z0-9_.]{2,})"',                          # double-quoted identifiers
        r"\b([A-Za-z_][A-Za-z0-9_]*(?:Parser|Handler|Manager|Client"
        r"|Server|Util|Helper|Builder|Factory|Engine|Reader|Writer"
        r"|Loader|Executor|Processor|Deserializer|Serializer))\b",   # class-suffix patterns
        r'\b(eval|exec|deserializ\w*|unpickl\w*|yaml\.load'
        r'|pickle\.load|subprocess|os\.system|Runtime\.exec'
        r'|ObjectInputStream|XStream|Kryo|Gson|Jackson'
        r'|XMLDecoder|Unmarshaller|Inflater|ZipInputStream'
        r'|ProcessBuilder|ScriptEngine|GroovyShell'
        r'|BeanUtils|PropertyUtils|OGNL|SpEL|EL expression'
        r'|ClassLoader|forName|loadClass|getMethod|invoke)\b',        # dangerous API names
        r'\b([A-Za-z]{3,}(?:Injection|Overflow|Traversal|Bypass'
        r'|Execution|Escalation|Forgery|Redirect))\b',               # vuln-type nouns
        r'(?:function|method|API|endpoint|feature|module|class|package)\s+'
        r'["\']?([A-Za-z_][A-Za-z0-9_.]{2,})["\']?',               # "function foo" etc.
        r'\b([A-Za-z]{3,}_[A-Za-z]{3,})\b',                        # underscore_identifiers
    ]

    # ── Helper functions ───────────────────────────────────────────────────────

    def extract_vuln_patterns(description: str) -> list:
        """Pull candidate identifiers out of a CVE description."""
        if not description:
            return []
        found = set()
        for pat in VULN_PATTERNS:
            for match in re.findall(pat, description, re.IGNORECASE):
                token = match if isinstance(match, str) else " ".join(match)
                token = token.strip("`\"' ").rstrip("()")
                if len(token) > 2 and not token.isdigit():
                    found.add(token.lower())
        return sorted(found)

    def load_zip_codebase(zip_bytes: bytes) -> dict:
        """Unzip and read all supported source files → {path: content}."""
        contents = {}
        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
            tmp.write(zip_bytes)
            tmp_path = tmp.name
        try:
            with zipfile.ZipFile(tmp_path, "r") as z:
                for name in z.namelist():
                    ext = os.path.splitext(name)[1].lower()
                    if ext in SUPPORTED_EXTS and not name.startswith("__MACOSX"):
                        try:
                            with z.open(name) as f:
                                contents[name] = f.read().decode("utf-8", errors="ignore")
                        except Exception:
                            pass
        finally:
            os.unlink(tmp_path)
        return contents

    def search_in_code(pattern: str, file_contents: dict) -> list:
        """Return list of {file, line, snippet} for every hit."""
        hits = []
        regex = re.compile(re.escape(pattern), re.IGNORECASE)
        for filepath, content in file_contents.items():
            for lineno, line in enumerate(content.splitlines(), 1):
                if regex.search(line):
                    hits.append({
                        "file":    filepath,
                        "line":    lineno,
                        "snippet": line.strip()[:200],
                    })
                    if len(hits) >= 50:   # cap total hits per pattern
                        return hits
        return hits

    def analyse_exposure(cve_list: list, file_contents: dict) -> list:
        """
        For each CVE dict {id, description, severity, component}:
          - Extract vulnerable patterns from description
          - Search codebase
          - Assign verdict: Exposed / Not Exposed / Cannot be Concluded
        """
        results = []
        for cve in cve_list:
            cve_id      = cve.get("id", "N/A")
            description = cve.get("description", "") or ""
            severity    = cve.get("severity", "UNKNOWN")
            component   = cve.get("component", "N/A")

            patterns = extract_vuln_patterns(description)

            if not patterns:
                results.append({
                    "CVE ID":            cve_id,
                    "Component":         component,
                    "Severity":          severity,
                    "Verdict":           "⚠️ Cannot be Concluded",
                    "Matched Patterns":  "",
                    "Description":       description[:300],
                    "_code_matches":     [],
                })
                continue

            all_hits        = []
            matched_pats    = []
            for pat in patterns:
                hits = search_in_code(pat, file_contents)
                if hits:
                    matched_pats.append(pat)
                    all_hits.extend(hits[:3])   # top 3 matches per pattern

            if all_hits:
                verdict = "🔴 Exposed / Used"
            else:
                verdict = "🟢 Not Exposed"

            results.append({
                "CVE ID":            cve_id,
                "Component":         component,
                "Severity":          severity,
                "Verdict":           verdict,
                "Matched Patterns":  ", ".join(matched_pats) if matched_pats else "—",
                "Description":       description[:300],
                "_code_matches":     all_hits,
            })
        return results

    def parse_scan_for_cves(scan_data: dict) -> list:
        """
        Extract CVE entries from CycloneDX JSON, generic SBOM JSON,
        or a plain list of {id, description, severity, component}.
        """
        cves = []

        # CycloneDX format
        if "vulnerabilities" in scan_data:
            for v in scan_data["vulnerabilities"]:
                vid  = v.get("id", "")
                desc = ""
                # description may be nested under advisories or detail
                if v.get("detail"):
                    desc = v["detail"]
                elif v.get("description"):
                    desc = v["description"]
                sev = "UNKNOWN"
                ratings = v.get("ratings", [])
                if ratings:
                    sev = ratings[0].get("severity", "UNKNOWN").upper()
                comp = ""
                affects = v.get("affects", [])
                if affects:
                    comp = affects[0].get("ref", "")
                cves.append({"id": vid, "description": desc, "severity": sev, "component": comp})

        # Generic flat list
        elif isinstance(scan_data, list):
            for item in scan_data:
                cves.append({
                    "id":          item.get("id", item.get("cve_id", item.get("CVE", ""))),
                    "description": item.get("description", item.get("desc", "")),
                    "severity":    item.get("severity", item.get("Severity", "UNKNOWN")),
                    "component":   item.get("component", item.get("package", "")),
                })

        # Nessus-style enriched dict (from Tab 2)
        elif "cve_enriched" in st.session_state:
            for cid, inf in st.session_state["cve_enriched"].items():
                desc = inf.get("nvd", {}).get("description") or inf.get("synopsis", "")
                cves.append({
                    "id":          cid,
                    "description": desc,
                    "severity":    inf.get("nessus_severity", "UNKNOWN"),
                    "component":   inf.get("plugin_name", ""),
                })

        return [c for c in cves if c.get("id")]

    # ── UI ─────────────────────────────────────────────────────────────────────
    st.markdown("""
<div class="status-info" style="margin-bottom:18px">
  <strong>🧬 Codebase Exposure Analysis</strong><br>
  Upload a vulnerability scan (CycloneDX JSON / SBOM) and your source code (ZIP).
  Each CVE description is parsed for vulnerable functions, classes, and API patterns,
  which are then searched in your codebase to determine real-world exposure.
</div>
""", unsafe_allow_html=True)

    # Upload row
    u1, u2 = st.columns(2)
    with u1:
        st.markdown('<div class="section-title">📄 Vulnerability Scan (JSON)</div>', unsafe_allow_html=True)
        scan_upload = st.file_uploader(
            "Upload CycloneDX / SBOM JSON",
            type=["json"],
            key="code_scan_upload",
            label_visibility="collapsed",
            help="CycloneDX BOM JSON, Grype/Trivy JSON output, or any flat CVE list JSON.",
        )
        st.caption("Supported: CycloneDX BOM, Grype, Trivy, or custom JSON with id/description/severity fields.")

        # Option: reuse Tab 2 data
        use_tab2 = False
        if "cve_enriched" in st.session_state:
            use_tab2 = st.checkbox(
                "♻️  Use CVE data already loaded in Tab 2 (CVE Analyzer)",
                key="code_use_tab2",
            )

    with u2:
        st.markdown('<div class="section-title">🗂️ Source Code (ZIP)</div>', unsafe_allow_html=True)
        code_upload = st.file_uploader(
            "Upload codebase ZIP",
            type=["zip"],
            key="code_zip_upload",
            label_visibility="collapsed",
            help="ZIP of your project. Scans: .py .js .ts .java .go .rb .php .cs .cpp .c .h .jsx .tsx .vue .scala .kt .rs .sh .yaml",
        )
        st.caption(f"Scanned extensions: {', '.join(sorted(SUPPORTED_EXTS))}")

    run_exposure = st.button(
        "▶  Run Exposure Analysis",
        type="primary",
        use_container_width=True,
        key="code_run",
    )

    if run_exposure:
        # ── Validate inputs ───────────────────────────────────────────────────
        if code_upload is None:
            st.error("⚠️  Please upload your source code as a ZIP file.")
            st.stop()
        if not use_tab2 and scan_upload is None:
            st.error("⚠️  Please upload a vulnerability scan JSON, or tick the Tab 2 reuse option.")
            st.stop()

        # ── Load codebase ─────────────────────────────────────────────────────
        with st.spinner("📦 Loading and indexing codebase…"):
            file_contents = load_zip_codebase(code_upload.read())

        if not file_contents:
            st.error("No supported source files found in the ZIP. Check the file contents.")
            st.stop()

        st.info(f"✅ Indexed **{len(file_contents)}** source files from ZIP")

        # ── Parse CVEs ────────────────────────────────────────────────────────
        with st.spinner("📋 Parsing vulnerability scan…"):
            if use_tab2:
                cve_list = parse_scan_for_cves({})
            else:
                try:
                    raw = json.loads(scan_upload.read().decode("utf-8", errors="ignore"))
                except json.JSONDecodeError as je:
                    st.error(f"Invalid JSON: {je}")
                    st.stop()
                cve_list = parse_scan_for_cves(raw)

        if not cve_list:
            st.error("No CVEs found in the scan. Ensure the JSON has 'vulnerabilities' or is a flat CVE list.")
            st.stop()

        st.info(f"✅ Found **{len(cve_list)}** CVEs to analyse")

        # ── Run analysis ──────────────────────────────────────────────────────
        prog = st.progress(0, text="Analysing CVE descriptions against codebase…")
        results = []
        for i, cve in enumerate(cve_list):
            prog.progress((i + 1) / len(cve_list), text=f"Checking {cve.get('id', '?')} ({i+1}/{len(cve_list)})…")
            results.extend(analyse_exposure([cve], file_contents))
        prog.empty()

        st.session_state["exposure_results"] = results
        st.success(f"✅ Analysis complete — {len(results)} CVEs processed")

    # ── Display results ────────────────────────────────────────────────────────
    if "exposure_results" in st.session_state:
        results = st.session_state["exposure_results"]

        exposed   = [r for r in results if r["Verdict"].startswith("🔴")]
        not_exp   = [r for r in results if r["Verdict"].startswith("🟢")]
        no_conc   = [r for r in results if r["Verdict"].startswith("⚠️")]

        # ── Summary metrics ────────────────────────────────────────────────────
        st.markdown('<div class="section-title">📊 Exposure Summary</div>', unsafe_allow_html=True)
        em1, em2, em3, em4 = st.columns(4)
        for col, val, color, label in [
            (em1, len(results),     "#e2e8f0", "Total CVEs"),
            (em2, len(exposed),     "#f87171", "🔴 Exposed / Used"),
            (em3, len(not_exp),     "#86efac", "🟢 Not Exposed"),
            (em4, len(no_conc),     "#fbbf24", "⚠️ Cannot be Concluded"),
        ]:
            with col:
                st.markdown(f"""
<div class="metric-card">
  <div class="val" style="color:{color}">{val}</div>
  <div class="lbl">{label}</div>
</div>""", unsafe_allow_html=True)

        st.markdown("&nbsp;", unsafe_allow_html=True)

        # ── Exposure donut chart ───────────────────────────────────────────────
        st.markdown('<div class="section-title">📈 Exposure Breakdown</div>', unsafe_allow_html=True)
        ch_a, ch_b = st.columns(2)

        with ch_a:
            fig_exp = px.pie(
                names=["Exposed / Used", "Not Exposed", "Cannot be Concluded"],
                values=[len(exposed), len(not_exp), len(no_conc)],
                color_discrete_sequence=["#dc2626", "#22c55e", "#d97706"],
                hole=0.55,
                title="CVE Exposure Verdict",
            )
            fig_exp.update_traces(textposition="outside", textinfo="percent+label")
            fig_exp.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0", title_font_size=13,
                showlegend=False, margin=dict(t=40, b=20, l=10, r=10),
            )
            st.plotly_chart(fig_exp, use_container_width=True)

        with ch_b:
            # Severity × Verdict stacked bar
            df_res = pd.DataFrame(results)
            if "Severity" in df_res.columns and not df_res.empty:
                df_res["Verdict_short"] = df_res["Verdict"].str.replace(r"^[^\w]+", "", regex=True).str.strip()
                sev_ord = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]
                df_res["Severity_up"] = df_res["Severity"].str.upper()
                grp = df_res.groupby(["Severity_up", "Verdict_short"]).size().reset_index(name="Count")
                fig_sv = px.bar(
                    grp, x="Severity_up", y="Count", color="Verdict_short",
                    color_discrete_map={
                        "Exposed / Used":        "#dc2626",
                        "Not Exposed":           "#22c55e",
                        "Cannot be Concluded":   "#d97706",
                    },
                    title="Severity × Exposure Verdict",
                    labels={"Severity_up": "Severity", "Verdict_short": "Verdict"},
                    category_orders={"Severity_up": sev_ord},
                    barmode="stack",
                )
                fig_sv.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", title_font_size=13,
                    xaxis=dict(showgrid=False),
                    yaxis=dict(showgrid=True, gridcolor="#334155"),
                    margin=dict(t=40, b=20, l=10, r=10),
                    legend=dict(font=dict(size=11)),
                )
                st.plotly_chart(fig_sv, use_container_width=True)

        # ── Filterable results table ───────────────────────────────────────────
        st.markdown('<div class="section-title">📋 CVE Exposure Results</div>', unsafe_allow_html=True)

        tf1, tf2, tf3 = st.columns([1, 1, 2])
        with tf1:
            verdict_opts = ["🔴 Exposed / Used", "🟢 Not Exposed", "⚠️ Cannot be Concluded"]
            verdict_filter = st.multiselect(
                "Filter by Verdict", verdict_opts, default=verdict_opts, key="exp_verdict_filter"
            )
        with tf2:
            sev_opts = sorted(set(r["Severity"] for r in results))
            sev_filter = st.multiselect(
                "Filter by Severity", sev_opts, default=sev_opts, key="exp_sev_filter"
            )
        with tf3:
            exp_search = st.text_input(
                "🔎  Search", placeholder="CVE ID, component, pattern…", key="exp_search"
            )

        df_display = pd.DataFrame([{
            "CVE ID":           r["CVE ID"],
            "Component":        r["Component"],
            "Severity":         r["Severity"],
            "Verdict":          r["Verdict"],
            "Matched Patterns": r["Matched Patterns"],
            "Description":      r["Description"][:150] + "…" if len(r["Description"]) > 150 else r["Description"],
        } for r in results])

        df_filtered = df_display[
            df_display["Verdict"].isin(verdict_filter) &
            df_display["Severity"].isin(sev_filter)
        ]
        if exp_search:
            mask = (
                df_filtered["CVE ID"].str.contains(exp_search, case=False, na=False) |
                df_filtered["Component"].str.contains(exp_search, case=False, na=False) |
                df_filtered["Matched Patterns"].str.contains(exp_search, case=False, na=False) |
                df_filtered["Description"].str.contains(exp_search, case=False, na=False)
            )
            df_filtered = df_filtered[mask]

        st.dataframe(df_filtered, use_container_width=True, height=400)
        st.caption(f"Showing {len(df_filtered)} of {len(results)} CVEs")

        # ── Exposed CVE deep dive ──────────────────────────────────────────────
        exposed_results = [r for r in results if r["Verdict"].startswith("🔴")]
        if exposed_results:
            st.markdown('<div class="section-title">🔎 Exposed CVE — Code Match Details</div>', unsafe_allow_html=True)

            sel_exp = st.selectbox(
                "Select an exposed CVE to inspect",
                options=[r["CVE ID"] for r in exposed_results],
                key="exp_cve_select",
                format_func=lambda c: next(
                    (f"{r['CVE ID']}  —  {r['Severity']}  |  Patterns: {r['Matched Patterns'][:60]}"
                     for r in exposed_results if r["CVE ID"] == c), c
                ),
            )

            if sel_exp:
                sel_r = next(r for r in exposed_results if r["CVE ID"] == sel_exp)
                st.markdown(f"""
<div style="background:#1e293b;border:1px solid #334155;border-radius:10px;padding:18px 20px;margin-bottom:14px">
  <div style="color:#38bdf8;font-size:0.78rem;font-weight:600;text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px">Description</div>
  <div style="color:#cbd5e1;font-size:0.87rem;line-height:1.6">{sel_r['Description']}</div>
  <div style="margin-top:12px;color:#fbbf24;font-size:0.8rem">
    <strong>Matched patterns:</strong> {sel_r['Matched Patterns']}
  </div>
</div>""", unsafe_allow_html=True)

                code_matches = sel_r.get("_code_matches", [])
                if code_matches:
                    for match in code_matches:
                        st.markdown(f"""
<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;padding:10px 14px;margin-bottom:8px;font-family:monospace;font-size:0.8rem">
  <span style="color:#64748b">{match['file']}</span>
  <span style="color:#475569"> · line {match['line']}</span><br>
  <span style="color:#aed581">{match['snippet']}</span>
</div>""", unsafe_allow_html=True)
                else:
                    st.info("No code snippets recorded for this CVE.")

        # ── Export ─────────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">💾 Export Exposure Report</div>', unsafe_allow_html=True)
        ex1, ex2 = st.columns(2)

        export_rows = [{
            "CVE ID":           r["CVE ID"],
            "Component":        r["Component"],
            "Severity":         r["Severity"],
            "Verdict":          r["Verdict"].replace("🔴 ", "").replace("🟢 ", "").replace("⚠️ ", ""),
            "Matched Patterns": r["Matched Patterns"],
            "Description":      r["Description"],
            "Code Matches":     "; ".join(
                f"{m['file']}:{m['line']}" for m in r.get("_code_matches", [])
            ),
        } for r in results]

        df_export = pd.DataFrame(export_rows)

        with ex1:
            buf_exp = BytesIO()
            with pd.ExcelWriter(buf_exp, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False, sheet_name="Exposure Report")
            st.download_button(
                "⬇  Download Excel",
                data=buf_exp.getvalue(),
                file_name=f"exposure_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with ex2:
            st.download_button(
                "⬇  Download CSV",
                data=df_export.to_csv(index=False).encode(),
                file_name=f"exposure_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    else:
        st.markdown("""
<div style="text-align:center;padding:60px 40px">
  <div style="font-size:3.5rem;margin-bottom:16px">🧬</div>
  <h3 style="color:#e2e8f0;margin:0 0 8px 0">No analysis run yet</h3>
  <p style="color:#64748b;margin:0">
    Upload a vulnerability scan JSON and your source code ZIP above,
    then click <strong>Run Exposure Analysis</strong>.
  </p>
</div>
""", unsafe_allow_html=True)
